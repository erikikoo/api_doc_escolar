from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import calendar
from pathlib import Path
from typing import Tuple
import base64
import io

# Configurações em português
MESES = {
    1: "JANEIRO", 2: "FEVEREIRO", 3: "MARÇO", 4: "ABRIL",
    5: "MAIO", 6: "JUNHO", 7: "JULHO", 8: "AGOSTO",
    9: "SETEMBRO", 10: "OUTUBRO", 11: "NOVEMBRO", 12: "DEZEMBRO"
}

DIAS_SEMANA = {
    "Monday": "SEGUNDA-FEIRA",
    "Tuesday": "TERÇA-FEIRA",
    "Wednesday": "QUARTA-FEIRA",
    "Thursday": "QUINTA-FEIRA",
    "Friday": "SEXTA-FEIRA"
}

AULAS = [f"{i}ª aula" for i in range(1, 10)]  # De 1ª até 9ª aula

def obter_dias_adjacentes(mes: int, ano: int) -> Tuple[list, list]:
    """Calcula dias dos meses adjacentes para completar semanas"""
    primeiro_dia = datetime(ano, mes, 1)
    ultimo_dia = datetime(ano, mes, calendar.monthrange(ano, mes)[1])
    
    dias_antes = []
    if primeiro_dia.weekday() > 0:  # Se não começa na segunda
        dias_antes = list(range(1, primeiro_dia.weekday() + 1))
    
    dias_depois = []
    if ultimo_dia.weekday() < 4:  # Se não termina na sexta
        dias_depois = list(range(1, 5 - ultimo_dia.weekday() + 1))
    
    return dias_antes, dias_depois

def criar_agenda(mes: int, ano: int, professor: str, return_base64: bool = True):
    """Gera a agenda e retorna o arquivo ou base64"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Agenda {mes:02d}-{ano}"

    # Configuração de estilos
    estilo_titulo = Font(bold=True, size=14)
    estilo_cabecalho = Font(bold=True, color="FFFFFF")
    estilo_dia = Font(bold=True)
    
    fill_titulo = PatternFill(start_color="D9E1F2", fill_type="solid")
    fill_cabecalho = PatternFill(start_color="4F81BD", fill_type="solid")
    fill_aula = PatternFill(start_color="F2F2F2", fill_type="solid")
    
    borda = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    alinhamento_centro = Alignment(horizontal="center", vertical="center")

    # Título principal
    ws.merge_cells('A1:F1')
    titulo = ws.cell(row=1, column=1, 
                    value=f"AGENDA PROFESSOR: {professor.upper()} - {MESES[mes]} {ano}")
    titulo.font = estilo_titulo
    titulo.fill = fill_titulo
    titulo.alignment = alinhamento_centro

    linha = 3  # Linha inicial para conteúdo
    dias_antes, dias_depois = obter_dias_adjacentes(mes, ano)

    for semana in calendar.monthcalendar(ano, mes):
        dias_semana_atual = [dia for dia in semana[:5] if dia != 0]
        primeira_semana = calendar.monthcalendar(ano, mes).index(semana) == 0
        ultima_semana = calendar.monthcalendar(ano, mes).index(semana) == len(calendar.monthcalendar(ano, mes)) - 1

        dias_completos = []
        if primeira_semana and dias_antes:
            dias_completos.extend([(None, 'anterior') for _ in dias_antes])
        dias_completos.extend([(dia, 'atual') for dia in dias_semana_atual])
        if ultima_semana and dias_depois:
            dias_completos.extend([(None, 'proximo') for _ in dias_depois])

        if not dias_completos:
            continue

        # Cabeçalhos
        ws.cell(row=linha, column=1, value="AULAS").font = estilo_cabecalho
        ws.cell(row=linha, column=1).fill = fill_cabecalho
        ws.cell(row=linha, column=1).border = borda
        ws.cell(row=linha, column=1).alignment = alinhamento_centro

        # Dias da semana
        col = 2
        for dia, tipo in dias_completos:
            celula = ws.cell(row=linha, column=col)
            celula.fill = fill_cabecalho
            celula.border = borda
            celula.alignment = alinhamento_centro

            if tipo == 'atual':
                data = datetime(ano, mes, dia)
                nome_dia = DIAS_SEMANA[data.strftime("%A")]
                celula.value = nome_dia
                celula.font = estilo_cabecalho

                celula_data = ws.cell(row=linha+1, column=col, value=f"{dia:02d}/{mes:02d}")
                celula_data.font = estilo_dia
                celula_data.border = borda
                celula_data.alignment = alinhamento_centro
            else:
                ws.cell(row=linha+1, column=col).border = borda
            col += 1

        # Grades das aulas
        for i, aula in enumerate(AULAS):
            celula_aula = ws.cell(row=linha+2+i, column=1, value=aula)
            celula_aula.font = Font()
            celula_aula.fill = fill_aula
            celula_aula.border = borda
            celula_aula.alignment = alinhamento_centro

            for col in range(2, len(dias_completos)+2):
                celula = ws.cell(row=linha+2+i, column=col)
                celula.border = borda
                if i == 0:
                    celula.fill = fill_aula

        linha += len(AULAS) + 2

    # Ajustes finais
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 18

    for row in range(1, linha):
        ws.row_dimensions[row].height = 20

    # Salvar em buffer ou arquivo
    if return_base64:
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        base64_data = base64.b64encode(buffer.getvalue()).decode("utf-8")
        return {
            "file_data": base64_data,
            "file_name": f"Agenda_{professor.replace(' ', '_')}_{mes:02d}_{ano}.xlsx",
            "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
    else:
        nome_arquivo = f"Agenda_{professor.replace(' ', '_')}_{mes:02d}_{ano}.xlsx"
        arquivo = Path(nome_arquivo)
        wb.save(arquivo)
        return arquivo